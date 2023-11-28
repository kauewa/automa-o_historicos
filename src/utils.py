import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import io
import zipfile

def create_zip_from_workbooks(workbooks):
    # Crie um buffer de bytes em memória para o arquivo zip
    zip_buffer = io.BytesIO()

    # Crie um objeto ZipFile que escreve no buffer de bytes
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        for wb in workbooks:
            # Crie um buffer de bytes em memória para o arquivo Excel
            excel_buffer = io.BytesIO()

            # Salve o Workbook neste buffer
            wb['dados'].save(excel_buffer)

            # Adicione o buffer do Excel ao arquivo Zip com um nome de arquivo
            zip_file.writestr(wb['nome_arquivo'], excel_buffer.getvalue())

    # Agora, zip_buffer contém os dados do arquivo zip
    return zip_buffer.getvalue()

def get_current_year():
    return datetime.now().year


def gerar_historicos_fund(df_alunos, df_sexto_ano, df_setimo_ano, df_oitavo_ano, df_nono_ano):
    # Lista para armazenar os históricos
    historicos = []
    
    # Iterar sobre cada aluno no dataframe de alunos

    for index, aluno in df_alunos.iterrows():
        notas_sexto_ano = df_sexto_ano[df_sexto_ano['Aluno'] == aluno['Nome do Aluno']]
        notas_setimo_ano = df_setimo_ano[df_setimo_ano['Aluno'] == aluno['Nome do Aluno']]
        notas_oitavo_ano = df_oitavo_ano[df_oitavo_ano['Aluno'] == aluno['Nome do Aluno']]
        notas_nono_ano = df_nono_ano[df_nono_ano['Aluno'] == aluno['Nome do Aluno']]
        data_nascimento = aluno["Data de Nascimento"]

        # Verificar se a data de nascimento não é uma string
        if not isinstance(data_nascimento, str):
            data_nascimento = data_nascimento.strftime('%d/%m/%Y')
        else:
            # Converter a string para um objeto datetime
            data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d %H:%M:%S')

            # Formatar a data no formato desejado
            data_nascimento = data_nascimento.strftime('%d/%m/%Y')

        # Atribuir cada nota a uma chave em um dicionário para o primeiro ano
        notas = {
            'nome': aluno["Nome do Aluno"],
            'ra': str(aluno["RA"]) + '-' + str(aluno["Dig. RA"]),
            'data_nascimento': data_nascimento,
            'portugues_sexto_ano': notas_sexto_ano['LINGUA PORTUGUESA'].values[0] if not notas_sexto_ano.empty else '',
            'ingles_sexto_ano': notas_sexto_ano['LINGUA INGLESA'].values[0] if not notas_sexto_ano.empty else '',
            'arte_sexto_ano': notas_sexto_ano['ARTE'].values[0] if not notas_sexto_ano.empty else '',
            'educacao_fisica_sexto_ano': notas_sexto_ano['EDUCACAO FISICA'].values[0] if not notas_sexto_ano.empty else '',
            'geografia_sexto_ano': notas_sexto_ano['GEOGRAFIA'].values[0] if not notas_sexto_ano.empty else '',
            'historia_sexto_ano': notas_sexto_ano['HISTORIA'].values[0] if not notas_sexto_ano.empty else '',
            'ciencias_sexto_ano': notas_sexto_ano['CIENCIAS'].values[0] if not notas_sexto_ano.empty else '',
            'matematica_sexto_ano': notas_sexto_ano['MATEMATICA'].values[0] if not notas_sexto_ano.empty else '',
            # Repita o processo para os outros anos
            'portugues_setimo_ano': notas_setimo_ano['LINGUA PORTUGUESA'].values[0] if not notas_setimo_ano.empty else '',
            'ingles_setimo_ano': notas_setimo_ano['LINGUA INGLESA'].values[0] if not notas_setimo_ano.empty else '',
            'arte_setimo_ano': notas_setimo_ano['ARTE'].values[0] if not notas_setimo_ano.empty else '',
            'educacao_fisica_setimo_ano': notas_setimo_ano['EDUCACAO FISICA'].values[0] if not notas_setimo_ano.empty else '',
            'geografia_setimo_ano': notas_setimo_ano['GEOGRAFIA'].values[0] if not notas_setimo_ano.empty else '',
            'historia_setimo_ano': notas_setimo_ano['HISTORIA'].values[0] if not notas_setimo_ano.empty else '',
            'ciencias_setimo_ano': notas_setimo_ano['CIENCIAS'].values[0] if not notas_setimo_ano.empty else '',
            'matematica_setimo_ano': notas_setimo_ano['MATEMATICA'].values[0] if not notas_setimo_ano.empty else '',
            'portugues_oitavo_ano': notas_oitavo_ano['LINGUA PORTUGUESA'].values[0] if not notas_oitavo_ano.empty else '',
            'ingles_oitavo_ano': notas_oitavo_ano['LINGUA INGLESA'].values[0] if not notas_oitavo_ano.empty else '',
            'arte_oitavo_ano': notas_oitavo_ano['ARTE'].values[0] if not notas_oitavo_ano.empty else '',
            'educacao_fisica_oitavo_ano': notas_oitavo_ano['EDUCACAO FISICA'].values[0] if not notas_oitavo_ano.empty else '',
            'geografia_oitavo_ano': notas_oitavo_ano['GEOGRAFIA'].values[0] if not notas_oitavo_ano.empty else '',
            'historia_oitavo_ano': notas_oitavo_ano['HISTORIA'].values[0] if not notas_oitavo_ano.empty else '',
            'ciencias_oitavo_ano': notas_oitavo_ano['CIENCIAS'].values[0] if not notas_oitavo_ano.empty else '',
            'matematica_oitavo_ano': notas_oitavo_ano['MATEMATICA'].values[0] if not notas_oitavo_ano.empty else '',
            'portugues_nono_ano': notas_nono_ano['LINGUA PORTUGUESA'].values[0] if not notas_nono_ano.empty else '',
            'ingles_nono_ano': notas_nono_ano['LINGUA INGLESA'].values[0] if not notas_nono_ano.empty else '',
            'arte_nono_ano': notas_nono_ano['ARTE'].values[0] if not notas_nono_ano.empty else '',
            'educacao_fisica_nono_ano': notas_nono_ano['EDUCACAO FISICA'].values[0] if not notas_nono_ano.empty else '',
            'geografia_nono_ano': notas_nono_ano['GEOGRAFIA'].values[0] if not notas_nono_ano.empty else '',
            'historia_nono_ano': notas_nono_ano['HISTORIA'].values[0] if not notas_nono_ano.empty else '',
            'ciencias_nono_ano': notas_nono_ano['CIENCIAS'].values[0] if not notas_nono_ano.empty else '',
            'matematica_nono_ano': notas_nono_ano['MATEMATICA'].values[0] if not notas_nono_ano.empty else '',
        }

        historico = preencher_modelo_fund(notas)
        historicos.append(historico)
    
    file_zip = create_zip_from_workbooks(historicos)
    return file_zip




def gerar_historicos(df_alunos, df_primeiro_ano, df_segundo_ano, df_terceiro_ano):
    # Lista para armazenar os históricos
    historicos = []
    
    # Iterar sobre cada aluno no dataframe de alunos

    for index, aluno in df_alunos.iterrows():
        notas_primeiro_ano = df_primeiro_ano[df_primeiro_ano['Aluno'] == aluno['Nome do Aluno']]
        notas_segundo_ano = df_segundo_ano[df_segundo_ano['Aluno'] == aluno['Nome do Aluno']]
        notas_terceiro_ano = df_terceiro_ano[df_terceiro_ano['Aluno'] == aluno['Nome do Aluno']]
        data_nascimento = aluno["Data de Nascimento"]

        # Verificar se a data de nascimento não é uma string
        if not isinstance(data_nascimento, str):
            data_nascimento = data_nascimento.strftime('%d/%m/%Y')
        else:
            # Converter a string para um objeto datetime
            data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d %H:%M:%S')

            # Formatar a data no formato desejado
            data_nascimento = data_nascimento.strftime('%d/%m/%Y')
        # Atribuir cada nota a uma chave em um dicionário para o primeiro ano
        notas = {
            'nome': aluno["Nome do Aluno"],
            'ra': str(aluno["RA"]) + '-' + str(aluno["Dig. RA"]),
            'data_nascimento': data_nascimento,
            'portugues_primeiro_ano': notas_primeiro_ano['LINGUA PORTUGUESA'].values[0] if not notas_primeiro_ano.empty else '',
            'ingles_primeiro_ano': notas_primeiro_ano['LINGUA INGLESA'].values[0] if not notas_primeiro_ano.empty else '',
            'arte_primeiro_ano': notas_primeiro_ano['ARTE'].values[0] if not notas_primeiro_ano.empty else '',
            'educacao_fisica_primeiro_ano': notas_primeiro_ano['EDUCACAO FISICA'].values[0] if not notas_primeiro_ano.empty else '',
            'geografia_primeiro_ano': notas_primeiro_ano['GEOGRAFIA'].values[0] if not notas_primeiro_ano.empty else '',
            'historia_primeiro_ano': notas_primeiro_ano['HISTORIA'].values[0] if not notas_primeiro_ano.empty else '',
            'sociologia_primeiro_ano': notas_primeiro_ano['SOCIOLOGIA'].values[0] if not notas_primeiro_ano.empty else '',
            'biologia_primeiro_ano': notas_primeiro_ano['BIOLOGIA'].values[0] if not notas_primeiro_ano.empty else '',
            'fisica_primeiro_ano': notas_primeiro_ano['FISICA'].values[0] if not notas_primeiro_ano.empty else '',
            'matematica_primeiro_ano': notas_primeiro_ano['MATEMATICA'].values[0] if not notas_primeiro_ano.empty else '',
            'quimica_primeiro_ano': notas_primeiro_ano['QUIMICA'].values[0] if not notas_primeiro_ano.empty else '',
            'filosofia_primeiro_ano': notas_primeiro_ano['FILOSOFIA'].values[0] if not notas_primeiro_ano.empty else '',
            'projeto_de_vida_primeiro_ano': notas_primeiro_ano['PROJETO DE VIDA'].values[0] if not notas_primeiro_ano.empty else '',
            # 'disciplinas_eletivas_primeiro_ano': notas_primeiro_ano['DISCIPLINAS ELETIVAS'].values[0] if not notas_primeiro_ano.empty else '',
            'tecnologia_e_inovacao_primeiro_ano': notas_primeiro_ano['TECNOLOGIA E INOVACAO'].values[0] if not notas_primeiro_ano.empty else '',
            'portugues_segundo_ano': notas_segundo_ano['LINGUA PORTUGUESA'].values[0] if not notas_segundo_ano.empty else '',
            'ingles_segundo_ano': notas_segundo_ano['LINGUA INGLESA'].values[0] if not notas_segundo_ano.empty else '',
            'arte_segundo_ano': notas_segundo_ano['ARTE'].values[0] if not notas_segundo_ano.empty else '',
            'educacao_fisica_segundo_ano': notas_segundo_ano['EDUCACAO FISICA'].values[0] if not notas_segundo_ano.empty else '',
            'geografia_segundo_ano': notas_segundo_ano['GEOGRAFIA'].values[0] if not notas_segundo_ano.empty else '',
            'historia_segundo_ano': notas_segundo_ano['HISTORIA'].values[0] if not notas_segundo_ano.empty else '',
            'sociologia_segundo_ano': notas_segundo_ano['SOCIOLOGIA'].values[0] if not notas_segundo_ano.empty else '',
            'biologia_segundo_ano': notas_segundo_ano['BIOLOGIA'].values[0] if not notas_segundo_ano.empty else '',
            'fisica_segundo_ano': notas_segundo_ano['FISICA'].values[0] if not notas_segundo_ano.empty else '',
            'matematica_segundo_ano': notas_segundo_ano['MATEMATICA'].values[0] if not notas_segundo_ano.empty else '',
            'quimica_segundo_ano': notas_segundo_ano['QUIMICA'].values[0] if not notas_segundo_ano.empty else '',
            'filosofia_segundo_ano': notas_segundo_ano['FILOSOFIA'].values[0] if not notas_segundo_ano.empty else '',
            'projeto_de_vida_segundo_ano': notas_segundo_ano['PROJETO DE VIDA'].values[0] if not notas_segundo_ano.empty else '',
            # 'disciplinas_eletivas_segundo_ano': notas_segundo_ano['DISCIPLINAS ELETIVAS'].values[0] if not notas_segundo_ano.empty else '',
            'tecnologia_e_inovacao_segundo_ano': notas_segundo_ano['TECNOLOGIA E INOVACAO'].values[0] if not notas_segundo_ano.empty else '',
            'portugues_terceiro_ano': notas_terceiro_ano['LINGUA PORTUGUESA'].values[0] if not notas_terceiro_ano.empty else '',
            'ingles_terceiro_ano': notas_terceiro_ano['LINGUA INGLESA'].values[0] if not notas_terceiro_ano.empty else '',
            'arte_terceiro_ano': notas_terceiro_ano['ARTE'].values[0] if not notas_terceiro_ano.empty else '',
            'educacao_fisica_terceiro_ano': notas_terceiro_ano['EDUCACAO FISICA'].values[0] if not notas_terceiro_ano.empty else '',
            'geografia_terceiro_ano': notas_terceiro_ano['GEOGRAFIA'].values[0] if not notas_terceiro_ano.empty else '',
            'historia_terceiro_ano': notas_terceiro_ano['HISTORIA'].values[0] if not notas_terceiro_ano.empty else '',
            'sociologia_terceiro_ano': notas_terceiro_ano['SOCIOLOGIA'].values[0] if not notas_terceiro_ano.empty else '',
            'biologia_terceiro_ano': notas_terceiro_ano['BIOLOGIA'].values[0] if not notas_terceiro_ano.empty else '',
            'fisica_terceiro_ano': notas_terceiro_ano['FISICA'].values[0] if not notas_terceiro_ano.empty else '',
            'matematica_terceiro_ano': notas_terceiro_ano['MATEMATICA'].values[0] if not notas_terceiro_ano.empty else '',
            'quimica_terceiro_ano': notas_terceiro_ano['QUIMICA'].values[0] if not notas_terceiro_ano.empty else '',
            'filosofia_terceiro_ano': notas_terceiro_ano['FILOSOFIA'].values[0] if not notas_terceiro_ano.empty else '',
            'projeto_de_vida_terceiro_ano': notas_terceiro_ano['PROJETO DE VIDA'].values[0] if not notas_terceiro_ano.empty else '',
            # 'disciplinas_eletivas_terceiro_ano': notas_terceiro_ano['DISCIPLINAS ELETIVAS'].values[0] if not notas_terceiro_ano.empty else '',
            'tecnologia_e_inovacao_terceiro_ano': notas_terceiro_ano['TECNOLOGIA E INOVACAO'].values[0] if not notas_terceiro_ano.empty else '',
        }


        historico = preencher_modelo(notas)
        historicos.append(historico)
    
    file_zip = create_zip_from_workbooks(historicos)
    return file_zip

def preencher_modelo_fund(dados):
    wb = load_workbook('modelo_historico_fund.xlsx')
    ws = wb['modelo']
    ws["B10"] = f"{dados["nome"]}"
    ws["L10"] = f"{dados["ra"]}"
    ws["D12"] = f"{dados["data_nascimento"]}"

    ws["K15"] = get_current_year() - 3
    ws["L15"] = get_current_year() - 2
    ws["M15"] = get_current_year() - 1
    ws["N15"] = get_current_year()
    ws["E40"] = get_current_year() - 3
    ws["E41"] = get_current_year() - 2
    ws["E42"] = get_current_year() - 1
    ws["E43"] = get_current_year()

    # Sexto ano
    ws["K17"] = dados['portugues_sexto_ano']
    ws["K18"] = dados['ingles_sexto_ano']
    ws["K28"] = dados['ingles_sexto_ano']
    ws["K19"] = dados['arte_sexto_ano']
    ws["K20"] = dados['educacao_fisica_sexto_ano']
    ws["K21"] = dados['historia_sexto_ano']
    ws["K22"] = dados['geografia_sexto_ano']
    ws["K23"] = dados['matematica_sexto_ano']
    ws["K24"] = dados['ciencias_sexto_ano']
    

    # Setimo ano
    ws["L17"] = dados['portugues_setimo_ano']
    ws["L18"] = dados['ingles_setimo_ano']
    ws["L28"] = dados['ingles_setimo_ano']
    ws["L19"] = dados['arte_setimo_ano']
    ws["L20"] = dados['educacao_fisica_setimo_ano']
    ws["L21"] = dados['historia_setimo_ano']
    ws["L22"] = dados['geografia_setimo_ano']
    ws["L23"] = dados['matematica_setimo_ano']
    ws["L24"] = dados['ciencias_setimo_ano']
    

    # Oitavo ano
    ws["M17"] = dados['portugues_oitavo_ano']
    ws["M18"] = dados['ingles_oitavo_ano']
    ws["M28"] = dados['ingles_oitavo_ano']
    ws["M19"] = dados['arte_oitavo_ano']
    ws["M20"] = dados['educacao_fisica_oitavo_ano']
    ws["M21"] = dados['historia_oitavo_ano']
    ws["M22"] = dados['geografia_oitavo_ano']
    ws["M23"] = dados['matematica_oitavo_ano']
    ws["M24"] = dados['ciencias_oitavo_ano']
    

    # Nono ano
    ws["N17"] = dados['portugues_nono_ano']
    ws["N18"] = dados['ingles_nono_ano']
    ws["N28"] = dados['ingles_nono_ano']
    ws["N19"] = dados['arte_nono_ano']
    ws["N20"] = dados['educacao_fisica_nono_ano']
    ws["N21"] = dados['historia_nono_ano']
    ws["N22"] = dados['geografia_nono_ano']
    ws["N23"] = dados['matematica_nono_ano']
    ws["N24"] = dados['ciencias_nono_ano']

    if type(ws["K17"].value) == int:
        ws["F40"] = '''PEI EE "PROFª MARIA ELISA DE OLIVEIRA"'''
        ws["K40"] = "São Miguel Arcanjo".upper()
        ws["K27"] = '1200'
        ws["K32"] = '320'
        ws["K33"] = '1520'
        ws["K29"] = 'F'
        ws["K30"] = 'F'
        ws["K31"] = 'F'
    if type(ws["L17"].value) == int:
        ws["F41"] = '''PEI EE "PROFª MARIA ELISA DE OLIVEIRA"'''
        ws["K41"] = "São Miguel Arcanjo".upper()
        ws["L27"] = '1200'
        ws["L32"] = '320'
        ws["L33"] = '1520'
        ws["L29"] = 'F'
        ws["L30"] = 'F'
        ws["L31"] = 'F'
    if type(ws["M17"].value) == int:
        ws["F42"] = '''PEI EE "PROFª MARIA ELISA DE OLIVEIRA"'''
        ws["K42"] = "São Miguel Arcanjo".upper()
        ws["M27"] = '1200'
        ws["M32"] = '320'
        ws["M33"] = '1520'
        ws["M29"] = 'F'
        ws["M30"] = 'F'
        ws["M31"] = 'F'
    if type(ws["N17"].value) == int:
        ws["F43"] = '''PEI EE "PROFª MARIA ELISA DE OLIVEIRA"'''
        ws["K43"] = "São Miguel Arcanjo".upper()
        ws["N27"] = '1200'
        ws["N32"] = '320'
        ws["N33"] = '1520'
        ws["N29"] = 'F'
        ws["N30"] = 'F'
        ws["N31"] = 'F'


    

    ws["A51"] = f'''O Diretor da PEIEE " Profª Maria Elisa de Oliveira", CERTIFICA, nos termos do Inciso VII, Artigo 24 da Lei Federal 9394/96, que {dados["nome"]} , RG: SP , concluiu o Ensino Fundamental, no ano de {get_current_year()}.'''

    return {
      'nome_arquivo': f'{dados['nome']}_fund.xlsx',
      'dados' : wb
    }

def preencher_modelo(dados):
    wb = load_workbook('modelo_historico.xlsx')
    ws = wb['modelo']
    ws["A10"] = f"Nome do Aluno: {dados["nome"]}"
    ws["R10"] = f"RA: {dados["ra"]}"
    ws["E12"] = f"Data: {dados["data_nascimento"]}"
    # Primeiro ano
    ws["P16"] = dados['portugues_primeiro_ano']
    ws["P17"] = dados['ingles_primeiro_ano']
    ws["P29"] = dados['ingles_primeiro_ano']
    ws["P18"] = dados['arte_primeiro_ano']
    ws["P19"] = dados['educacao_fisica_primeiro_ano']
    ws["P25"] = dados['geografia_primeiro_ano']
    ws["P24"] = dados['historia_primeiro_ano']
    ws["P27"] = dados['sociologia_primeiro_ano']
    ws["P21"] = dados['biologia_primeiro_ano']
    ws["P22"] = dados['fisica_primeiro_ano']
    ws["P20"] = dados['matematica_primeiro_ano']
    ws["P23"] = dados['quimica_primeiro_ano']
    ws["P26"] = dados['filosofia_primeiro_ano']

    # Segundo ano
    ws["R16"] = dados['portugues_segundo_ano']
    ws["R17"] = dados['ingles_segundo_ano']
    ws["R29"] = dados['ingles_segundo_ano']
    ws["R18"] = dados['arte_segundo_ano']
    ws["R19"] = dados['educacao_fisica_segundo_ano']
    ws["R25"] = dados['geografia_segundo_ano']
    ws["R24"] = dados['historia_segundo_ano']
    ws["R27"] = dados['sociologia_segundo_ano']
    ws["R21"] = dados['biologia_segundo_ano']
    ws["R22"] = dados['fisica_segundo_ano']
    ws["R20"] = dados['matematica_segundo_ano']
    ws["R23"] = dados['quimica_segundo_ano']
    ws["R26"] = dados['filosofia_segundo_ano']

    # Terceiro ano
    ws["T16"] = dados['portugues_terceiro_ano']
    ws["T17"] = dados['ingles_terceiro_ano']
    ws["T29"] = dados['ingles_terceiro_ano']
    ws["T18"] = dados['arte_terceiro_ano']
    ws["T19"] = dados['educacao_fisica_terceiro_ano']
    ws["T25"] = dados['geografia_terceiro_ano']
    ws["T24"] = dados['historia_terceiro_ano']
    ws["T27"] = dados['sociologia_terceiro_ano']
    ws["T21"] = dados['biologia_terceiro_ano']
    ws["T22"] = dados['fisica_terceiro_ano']
    ws["T20"] = dados['matematica_terceiro_ano']
    ws["T23"] = dados['quimica_terceiro_ano']
    ws["T26"] = dados['filosofia_terceiro_ano']

    ws["P33"] = '-'
    ws["P30"] = '-'
    ws["P31"] = '-'
    ws["R33"] = '-'
    ws["R30"] = '-'
    ws["R31"] = '-'
    ws["T33"] = '-'
    ws["T30"] = '-'
    ws["T31"] = '-'

    ws["G43"] = get_current_year() - 3
    ws["G45"] = get_current_year() - 2
    ws["G46"] = get_current_year() - 1
    ws["G47"] = get_current_year()

    if type(ws["P16"].value) == int:
        ws["I45"] = '''PEI EE "PROFª MARIA ELISA DE OLIVEIRA"'''
        ws["Q45"] = "São Miguel Arcanjo/SP"
        ws["P28"] = '1200'
        ws["P39"] = '320'
        ws["P41"] = '1520'
        ws["P33"] = 'F'
        ws["P30"] = 'F'
        ws["P31"] = 'F'
    if type(ws["R16"].value) == int:
        ws["I46"] = '''PEI EE "PROFª MARIA ELISA DE OLIVEIRA"'''
        ws["Q46"] = "São Miguel Arcanjo/SP"
        ws["R28"] = '1200'
        ws["R39"] = '320'
        ws["R41"] = '1520'
        ws["R33"] = 'F'
        ws["R30"] = 'F'
        ws["R31"] = 'F'
    if type(ws["T16"].value) == int:
        ws["I47"] = '''PEI EE "PROFª MARIA ELISA DE OLIVEIRA"'''
        ws["Q47"] = "São Miguel Arcanjo/SP"
        ws["T28"] = '720'
        ws["T39"] = '800'
        ws["T41"] = '1520'
        ws["T33"] = 'F'
        ws["T30"] = 'F'
        ws["T31"] = 'F'


    ws["A54"] = f'''O Diretor da PEI EE "PROFª MARIA ELISA DE OLIVEIRA", CERTIFICA, nos termos do Inciso VII, Artigo 24 da Lei Federal 9394/96, que {dados["nome"]}, RG  SP, concluiu o Ensino Médio, no ano de {get_current_year()}.'''


    return {
      'nome_arquivo': f'{dados["nome"]}.xlsx',
      'dados' : wb
    }

